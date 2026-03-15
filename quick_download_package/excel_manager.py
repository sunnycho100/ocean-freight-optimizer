"""
Excel File Management Module
Handles saving data to Excel files with versioning and appending logic.
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelManager:
    """Manages Excel file operations for scraped data."""
    
    def __init__(self, output_dir=None, error_folder=None):
        """
        Initialize Excel manager.
        
        Args:
            output_dir: Directory for output files (defaults to ./downloads)
            error_folder: Directory for error logs (defaults to ./scraping_errors)
        """
        self.output_dir = output_dir or os.path.join(os.getcwd(), "downloads")
        self.error_folder = error_folder or os.path.join(os.getcwd(), "scraping_errors")
        self.current_run_filepath = None
        self.first_save_done = False
        
        # Ensure directories exist
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        if not os.path.exists(self.error_folder):
            os.makedirs(self.error_folder)
    
    def save_to_excel(self, df, filename, destination):
        """
        Save DataFrame to Excel file.
        
        Handles:
        - Creating new versioned file if one from previous run exists
        - Appending to same file during current run
        
        Args:
            df: pandas.DataFrame to save
            filename: Name of the Excel file
            destination: Destination name (for logging)
            
        Returns:
            bool: True if save successful, False otherwise
        """
        print(f"\n>>> [SAVING] Saving to Excel file: {filename}")
        
        if df is None or len(df) == 0:
            print("   [WARNING] No data to save!")
            self._log_no_data_error(destination)
            return False
        
        try:
            if not self.first_save_done:
                filepath = self._prepare_first_save_path(filename)
                success = self._write_excel(df, filepath)
                if success:
                    self.current_run_filepath = filepath
                    self.first_save_done = True
                return success
            else:
                return self._append_to_excel(df, self.current_run_filepath)
            
        except Exception as e:
            print(f"   [ERROR] Failed to save Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _prepare_first_save_path(self, filename):
        """
        Determine the filepath for the first save in this run.
        If today's base filename already exists, create a versioned filename.
        
        Args:
            filename: Base filename
            
        Returns:
            str: Path to use for saving
        """
        filepath = os.path.join(self.output_dir, filename)
        if os.path.exists(filepath):
            print(f"   [Info] File from previous run exists - creating new version...")
            return self._create_versioned_filename(filename)
        else:
            print(f"   [Info] Creating new file for this run...")
            return filepath
    
    def _create_versioned_filename(self, filename):
        """
        Create a versioned filename (_2, _3, etc.).
        
        Args:
            filename: Original filename
            
        Returns:
            str: Versioned file path
        """
        name, ext = os.path.splitext(filename)
        version = 2
        while os.path.exists(os.path.join(self.output_dir, f"{name}_{version}{ext}")):
            version += 1
        
        versioned_filepath = os.path.join(self.output_dir, f"{name}_{version}{ext}")
        filename_used = f"{name}_{version}{ext}"
        print(f"   [Info] Using: {filename_used}")
        
        return versioned_filepath
    
    def _write_excel(self, df, filepath):
        """
        Write DataFrame to a new Excel file.
        
        Args:
            df: DataFrame to write
            filepath: Full path to the file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            df.to_excel(filepath, index=False, sheet_name="Inland Rates")
            print(f"   [SUCCESS] Saved {len(df)} rows to: {filepath}")
            print(f"   [Info] Columns: {list(df.columns)}")
            return True
        except PermissionError:
            print(f"   [ERROR] File is open in another program. Please close it and try again.")
            return False
        except Exception as e:
            print(f"   [ERROR] Error writing Excel file: {e}")
            return False

    def _append_to_excel(self, df, filepath):
        """
        Append new rows to an existing Excel file without re-reading all data.

        Args:
            df: DataFrame to append
            filepath: Existing Excel file path

        Returns:
            bool: True if append succeeded, False otherwise
        """
        try:
            print(f"   [Info] Appending to current run's file: {os.path.basename(filepath)}")

            wb = load_workbook(filepath)
            ws = wb["Inland Rates"] if "Inland Rates" in wb.sheetnames else wb.active

            before_rows = max(ws.max_row - 1, 0)  # subtract header
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)
            after_rows = max(ws.max_row - 1, 0)

            wb.save(filepath)

            print(
                f"   [SUCCESS] Appended {len(df)} rows "
                f"(previous: {before_rows}, total: {after_rows}) to: {filepath}"
            )
            return True

        except PermissionError:
            print(f"   [ERROR] File is open in another program. Please close it and try again.")
            return False
        except Exception as e:
            print(f"   [ERROR] Error appending Excel file: {e}")
            return False
    
    def _log_no_data_error(self, destination):
        """
        Log an error when no data is available to save.
        
        Args:
            destination: Destination name
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest_clean = destination.replace(", ", "_").replace(" ", "_")
            error_msg = (f"No data to save for {destination}\n"
                        f"Timestamp: {timestamp}\n"
                        f"Reason: Empty data returned from scraper")
            log_path = os.path.join(self.error_folder, f"NO_DATA_{dest_clean}_{timestamp}.txt")
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(error_msg)
            print(f"   [ERROR] Error log saved: {log_path}")
        except Exception as log_err:
            print(f"   [WARNING] Could not save error log: {log_err}")
    
    def save_exception_log(self, destination, error, driver=None):
        """
        Save error log and screenshot when exception occurs.
        
        Args:
            destination: Destination name
            error: Exception object
            driver: WebDriver instance (optional, for screenshot)
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest_clean = destination.replace(", ", "_").replace(" ", "_")
            
            # Save screenshot if driver available
            if driver:
                screenshot_path = os.path.join(self.error_folder, 
                                             f"EXCEPTION_{dest_clean}_{timestamp}.png")
                driver.save_screenshot(screenshot_path)
                print(f"   [ERROR] Screenshot saved: {screenshot_path}")
            
            # Save error log
            error_log = (f"Exception processing: {destination}\n"
                        f"Timestamp: {timestamp}\n"
                        f"Error: {error}\n\n"
                        f"Stack trace available in console output")
            log_path = os.path.join(self.error_folder, 
                                   f"EXCEPTION_{dest_clean}_{timestamp}.txt")
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(error_log)
            print(f"   [ERROR] Error log saved: {log_path}")
            
        except Exception as save_err:
            print(f"   [WARNING] Could not save error files: {save_err}")
    
    def reset_for_new_run(self):
        """Reset state for a new run."""
        self.current_run_filepath = None
        self.first_save_done = False
