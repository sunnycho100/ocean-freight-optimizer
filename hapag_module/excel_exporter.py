"""
Excel export functionality for Hapag-Lloyd surcharge data.

This module handles Excel file creation, formatting, and data export
with smart column mapping and file management.
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelExporter:
    """Handles Excel file creation and data export."""
    
    def __init__(self, base_dir: str = None):
        """
        Initialize ExcelExporter.
        
        Args:
            base_dir: Base directory for downloads. Defaults to current working directory.
        """
        self.base_dir = base_dir or os.getcwd()
        self.downloads_dir = os.path.join(self.base_dir, "downloads")
        self.fixed_headers = ["From", "To", "Via", "Description", "Curr.", "20STD", "40STD", "40HC", "Transport Remarks"]
        
        # Ensure downloads directory exists
        os.makedirs(self.downloads_dir, exist_ok=True)
    
    def determine_excel_filename(self, base_filename: str = "hapag_surcharges.xlsx") -> str:
        """
        Determine the Excel filename to use for this run.
        Checks once at the start if base file exists, creates dated version if needed.
        
        Args:
            base_filename: Base filename to use
            
        Returns:
            Filename to use (e.g., "hapag_surcharges.xlsx" or "hapag_surcharges_20260113_2.xlsx")
        """
        base_path = os.path.join(self.downloads_dir, base_filename)
        
        if not os.path.exists(base_path):
            return base_filename
        
        # Create dated filename
        today = datetime.now().strftime("%Y%m%d")
        base_name = base_filename.rsplit('.', 1)[0]
        ext = base_filename.rsplit('.', 1)[1] if '.' in base_filename else 'xlsx'
        
        # Check if dated file exists, add _2, _3, etc.
        counter = 1
        dated_filename = f"{base_name}_{today}.{ext}"
        dated_path = os.path.join(self.downloads_dir, dated_filename)
        
        while os.path.exists(dated_path):
            counter += 1
            dated_filename = f"{base_name}_{today}_{counter}.{ext}"
            dated_path = os.path.join(self.downloads_dir, dated_filename)
        
        return dated_filename
    
    def _create_new_workbook(self, excel_path: str) -> tuple[Workbook, Any]:
        """
        Create a new Excel workbook with headers and metadata.
        
        Args:
            excel_path: Path where file will be saved
            
        Returns:
            Tuple of (workbook, worksheet)
        """
        print(f"   [INFO] Creating new file: {excel_path}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Surcharges"
        
        # Add metadata
        ws["H1"] = "Generated:"
        ws["I1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["H2"] = "Origin:"
        ws["I2"] = "BUSAN (KRPUS)"
        
        # Write headers with formatting
        for col_idx, header in enumerate(self.fixed_headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return wb, ws
    
    def _load_existing_workbook(self, excel_path: str) -> tuple[Workbook, Any]:
        """
        Load existing Excel workbook for appending data.
        
        Args:
            excel_path: Path to existing file
            
        Returns:
            Tuple of (workbook, worksheet)
        """
        print(f"   [INFO] Appending to existing file: {excel_path}")
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        return wb, ws
    
    def _map_container_values(self, row_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Map container values based on actual columns present in source table.
        Values are mapped by column name, not position.
        
        Args:
            row_data: Row data dictionary
            
        Returns:
            Dictionary with mapped values for 20STD, 40STD, 40HC
        """
        containers = row_data["container_values"]
        
        # Get values directly from containers - only use what's actually in the source
        val_20std = containers.get("20STD", "")
        val_40std = containers.get("40STD", "")
        val_40hc = containers.get("40HC", "")
        
        return {
            "20STD": val_20std,
            "40STD": val_40std,
            "40HC": val_40hc
        }
    
    def _write_data_rows(self, ws: Any, data: List[Dict[str, Any]], route_info: Dict[str, str], start_row: int) -> int:
        """
        Write data rows to worksheet.
        
        Args:
            ws: Worksheet object
            data: List of data dictionaries
            route_info: Route information dictionary
            start_row: Starting row number
            
        Returns:
            Next available row number
        """
        current_row = start_row
        
        for row_data in data:
            desc = row_data["description"]
            curr = row_data["curr"]
            remarks = row_data["remarks"]
            
            # Map container values
            container_values = self._map_container_values(row_data)
            
            # Create row values matching fixed headers
            row_values = [
                route_info.get("from", ""),
                route_info.get("to", ""),
                route_info.get("via", ""),
                desc,
                curr,
                container_values["20STD"],
                container_values["40STD"],
                container_values["40HC"],
                remarks
            ]
            
            # Write to worksheet
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
        
        return current_row
    
    def _adjust_column_widths(self, ws: Any) -> None:
        """
        Adjust column widths for better readability.
        
        Args:
            ws: Worksheet object
        """
        ws.column_dimensions['A'].width = 20  # From
        ws.column_dimensions['B'].width = 25  # To
        ws.column_dimensions['C'].width = 20  # Via
        ws.column_dimensions['D'].width = 40  # Description
        ws.column_dimensions['E'].width = 10  # Curr.
        ws.column_dimensions['F'].width = 12  # 20STD
        ws.column_dimensions['G'].width = 12  # 40STD
        ws.column_dimensions['H'].width = 12  # 40HC
        ws.column_dimensions['I'].width = 50  # Transport Remarks
    
    def save_to_excel(self, data: List[Dict[str, Any]], destination: str = "", 
                     route_info: Optional[Dict[str, str]] = None, 
                     filename: str = "hapag_surcharges.xlsx") -> str:
        """
        Save extracted data to Excel file with smart column mapping.
        - Fixed columns: From, To, Via, Description, Curr., 20STD, 40STD, 40HC, Transport Remarks
        - If 20STD not in data, copy 40STD value
        - Appends to existing file if present
        
        Args:
            data: List of dicts with keys: description, curr, container_values, remarks
            destination: Name of the destination city/country
            route_info: Dict with keys: from, to, via
            filename: Name of the Excel file
            
        Returns:
            Path to saved Excel file
        """
        if route_info is None:
            route_info = {"from": "", "to": "", "via": ""}
        
        excel_path = os.path.join(self.downloads_dir, filename)
        
        # Check if file exists for append mode
        if os.path.exists(excel_path):
            wb, ws = self._load_existing_workbook(excel_path)
            start_row = ws.max_row + 1
        else:
            wb, ws = self._create_new_workbook(excel_path)
            start_row = 5  # First data row after headers
        
        # Write data rows
        next_row = self._write_data_rows(ws, data, route_info, start_row)
        
        # Adjust column widths
        self._adjust_column_widths(ws)
        
        # Save workbook
        try:
            wb.save(excel_path)
            print(f"✅ Data saved successfully to: {excel_path}")
            return excel_path
            
        except Exception as e:
            print(f"❌ ERROR saving Excel file: {e}")
            raise
    
    def get_excel_path(self, filename: str) -> str:
        """
        Get full path to Excel file in downloads directory.
        
        Args:
            filename: Excel filename
            
        Returns:
            Full path to Excel file
        """
        return os.path.join(self.downloads_dir, filename)
    
    def file_exists(self, filename: str) -> bool:
        """
        Check if Excel file exists in downloads directory.
        
        Args:
            filename: Excel filename
            
        Returns:
            True if file exists, False otherwise
        """
        return os.path.exists(self.get_excel_path(filename))
    
    def get_downloads_dir(self) -> str:
        """
        Get downloads directory path.
        
        Returns:
            Path to downloads directory
        """
        return self.downloads_dir
    
    def validate_data_for_export(self, data: List[Dict[str, Any]]) -> bool:
        """
        Validate data before export.
        
        Args:
            data: List of data dictionaries
            
        Returns:
            True if data is valid for export, False otherwise
        """
        if not data:
            print("   [EXPORT VALIDATION] ❌ No data to export")
            return False
        
        valid_count = 0
        for item in data:
            if (item.get("description") and 
                item.get("container_values") and
                len(item["container_values"]) > 0):
                valid_count += 1
        
        if valid_count == 0:
            print("   [EXPORT VALIDATION] ❌ No valid rows found for export")
            return False
        
        print(f"   [EXPORT VALIDATION] ✅ {valid_count}/{len(data)} rows ready for export")
        return True