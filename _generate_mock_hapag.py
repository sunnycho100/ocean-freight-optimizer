"""
Generate mock hapag_surcharges.xlsx file based on the HAPAG data structure
documented in dataflow.md, with rates benchmarked from ONE Inland portal data.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_mock_hapag_data():
    """Create mock HAPAG surcharges data for all destinations."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Surcharges"
    
    # Row 1-3: Metadata (these are the "4 skipped header rows" - rows 1-3 are metadata, row 4 is headers)
    ws["H1"] = "Generated:"
    ws["I1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["H2"] = "Origin:"
    ws["I2"] = "BUSAN (KRPUS)"
    ws["H3"] = "Mock Data"
    ws["I3"] = "Benchmarked from ONE Inland Portal"
    
    # Row 4: Column headers
    headers = ["From", "To", "Via", "Description", "Curr.", "20STD", "40STD", "40HC", "Transport Remarks"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Define mock data for all destinations
    # Rates benchmarked from ONE Inland data with realistic HAPAG pricing
    # HAPAG uses USD for ocean freight and EUR for inland/surcharges
    destinations_data = [
        # --- VALENCE, DROME, FRANCE ---
        {
            "to": "VALENCE, DROME",
            "via": "FOS SUR MER",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1750, "40": 2850, "40hc": 2950},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; FOS SUR MER - VALENCE", "curr": "EUR", "20": 450, "40": 650, "40hc": 650},
                     {"desc": "From FOS SUR MER; via Truck", "curr": "EUR", "20": 580, "40": 780, "40hc": 780},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 185, "40": 280, "40hc": 280},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- ARQUES-LA-BATAILLE, FRANCE ---
        {
            "to": "ARQUES-LA-BATAILLE",
            "via": "LE HAVRE",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1720, "40": 2800, "40hc": 2900},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "From LE HAVRE; via Truck", "curr": "EUR", "20": 380, "40": 520, "40hc": 520},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 185, "40": 280, "40hc": 280},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- LEUTKIRCH IM ALLGAEU, BW, GERMANY ---
        {
            "to": "LEUTKIRCH IM ALLGAEU",
            "via": "HAMBURG",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; HAMBURG - LEUTKIRCH IM ALLGAEU", "curr": "EUR", "20": 820, "40": 1050, "40hc": 1050},
                     {"desc": "From HAMBURG; via Truck", "curr": "EUR", "20": 950, "40": 1280, "40hc": 1280},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- FUERTH, BY, GERMANY ---
        {
            "to": "FUERTH",
            "via": "HAMBURG",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; HAMBURG - FUERTH", "curr": "EUR", "20": 680, "40": 920, "40hc": 920},
                     {"desc": "From HAMBURG; via Truck", "curr": "EUR", "20": 810, "40": 1100, "40hc": 1100},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- FORCHHEIM, BY, GERMANY ---
        {
            "to": "FORCHHEIM",
            "via": "HAMBURG",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; HAMBURG - FORCHHEIM", "curr": "EUR", "20": 710, "40": 960, "40hc": 960},
                     {"desc": "From HAMBURG; via Truck", "curr": "EUR", "20": 840, "40": 1140, "40hc": 1140},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- MUENSTER, NW, GERMANY ---
        {
            "to": "MUENSTER",
            "via": "HAMBURG",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; HAMBURG - MUENSTER", "curr": "EUR", "20": 580, "40": 780, "40hc": 780},
                     {"desc": "From HAMBURG; via Truck", "curr": "EUR", "20": 720, "40": 980, "40hc": 980},
                     {"desc": "From BREMERHAVEN; via Truck", "curr": "EUR", "20": 650, "40": 880, "40hc": 880},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- ORANIENBURG, BB, GERMANY ---
        {
            "to": "ORANIENBURG",
            "via": "HAMBURG",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; HAMBURG - ORANIENBURG", "curr": "EUR", "20": 490, "40": 680, "40hc": 680},
                     {"desc": "From HAMBURG; via Truck", "curr": "EUR", "20": 620, "40": 850, "40hc": 850},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- WEITERSTADT, HE, GERMANY ---
        {
            "to": "WEITERSTADT",
            "via": "ROTTERDAM",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Barge/Truck; ROTTERDAM - WEITERSTADT", "curr": "EUR", "20": 560, "40": 760, "40hc": 760},
                     {"desc": "From ROTTERDAM; via Truck", "curr": "EUR", "20": 690, "40": 940, "40hc": 940},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- DORTMUND, NW, GERMANY ---
        {
            "to": "DORTMUND",
            "via": "HAMBURG",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; HAMBURG - DORTMUND", "curr": "EUR", "20": 540, "40": 730, "40hc": 730},
                     {"desc": "From HAMBURG; via Truck", "curr": "EUR", "20": 680, "40": 920, "40hc": 920},
                     {"desc": "From BREMERHAVEN; via Truck", "curr": "EUR", "20": 610, "40": 830, "40hc": 830},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- WORMS, RP, GERMANY ---
        {
            "to": "WORMS",
            "via": "ROTTERDAM",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Barge/Truck; ROTTERDAM - WORMS", "curr": "EUR", "20": 590, "40": 800, "40hc": 800},
                     {"desc": "From ROTTERDAM; via Truck", "curr": "EUR", "20": 730, "40": 990, "40hc": 990},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- DIETZENBACH, HE, GERMANY ---
        {
            "to": "DIETZENBACH",
            "via": "ROTTERDAM",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Barge/Truck; ROTTERDAM - DIETZENBACH", "curr": "EUR", "20": 570, "40": 770, "40hc": 770},
                     {"desc": "From ROTTERDAM; via Truck", "curr": "EUR", "20": 700, "40": 950, "40hc": 950},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- TEVEROLA, ITALY ---
        {
            "to": "TEVEROLA",
            "via": "NAPLES",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1950, "40": 3200, "40hc": 3300},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "From NAPLES; via Truck", "curr": "EUR", "20": 280, "40": 380, "40hc": 380},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 195, "40": 295, "40hc": 295},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 165, "40": 330, "40hc": 330},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- OEGSTGEEST, NETHERLANDS ---
        {
            "to": "OEGSTGEEST",
            "via": "ROTTERDAM",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "From ROTTERDAM; via Truck", "curr": "EUR", "20": 350, "40": 480, "40hc": 480},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- KOBIERZYCE, POLAND ---
        {
            "to": "KOBIERZYCE",
            "via": "GDANSK",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1780, "40": 2900, "40hc": 3000},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; GDANSK - KOBIERZYCE", "curr": "EUR", "20": 480, "40": 720, "40hc": 720},
                     {"desc": "From GDANSK; via Truck", "curr": "EUR", "20": 620, "40": 880, "40hc": 880},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 165, "40": 250, "40hc": 250},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- KOMAROM, HUNGARY ---
        {
            "to": "KOMAROM",
            "via": "KOPER",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1850, "40": 3050, "40hc": 3150},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; KOPER - KOMAROM", "curr": "EUR", "20": 620, "40": 890, "40hc": 890},
                     {"desc": "From KOPER; via Truck", "curr": "EUR", "20": 750, "40": 1050, "40hc": 1050},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 180, "40": 270, "40hc": 270},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 155, "40": 310, "40hc": 310},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- BAUMGARTENBERG, AUSTRIA ---
        {
            "to": "BAUMGARTENBERG",
            "via": "HAMBURG",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1680, "40": 2750, "40hc": 2850},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "Combined Road/Rail; HAMBURG - BAUMGARTENBERG", "curr": "EUR", "20": 880, "40": 1150, "40hc": 1150},
                     {"desc": "From HAMBURG; via Truck", "curr": "EUR", "20": 1050, "40": 1380, "40hc": 1380},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 175, "40": 265, "40hc": 265},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 148, "40": 296, "40hc": 296},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
        # --- TAMPERE, FINLAND ---
        {
            "to": "TAMPERE",
            "via": "HELSINKI",
            "charges": [
                {"desc": "Ocean Freight Basic", "curr": "USD", "20": 1900, "40": 3100, "40hc": 3200},
                {"desc": "Destination Landfreight", "curr": "", "20": "", "40": "", "40hc": "",
                 "sub_options": [
                     {"desc": "From HELSINKI; via Truck", "curr": "EUR", "20": 420, "40": 590, "40hc": 590},
                     {"desc": "From KOTKA; via Truck", "curr": "EUR", "20": 480, "40": 650, "40hc": 650},
                 ]},
                {"desc": "Terminal Handling Charge Dest.", "curr": "EUR", "20": 190, "40": 285, "40hc": 285},
                {"desc": "ISPS Surcharge", "curr": "USD", "20": 10, "40": 10, "40hc": 10},
                {"desc": "Low Sulphur Surcharge", "curr": "USD", "20": 160, "40": 320, "40hc": 320},
                {"desc": "Seal Fee", "curr": "USD", "20": 15, "40": 15, "40hc": 15},
            ]
        },
    ]
    
    # Write data rows starting at row 5
    current_row = 5
    
    for dest_data in destinations_data:
        to_val = dest_data["to"]
        via_val = dest_data["via"]
        
        for charge in dest_data["charges"]:
            desc = charge["desc"]
            curr = charge["curr"]
            val_20 = charge["20"]
            val_40 = charge["40"]
            val_40hc = charge["40hc"]
            
            # Write main charge row
            ws.cell(row=current_row, column=1, value="BUSAN")
            ws.cell(row=current_row, column=2, value=to_val)
            ws.cell(row=current_row, column=3, value=via_val)
            ws.cell(row=current_row, column=4, value=desc)
            ws.cell(row=current_row, column=5, value=curr)
            ws.cell(row=current_row, column=6, value=val_20)
            ws.cell(row=current_row, column=7, value=val_40)
            ws.cell(row=current_row, column=8, value=val_40hc)
            ws.cell(row=current_row, column=9, value="")
            current_row += 1
            
            # Write sub-options if they exist
            if "sub_options" in charge:
                for sub in charge["sub_options"]:
                    ws.cell(row=current_row, column=1, value="BUSAN")
                    ws.cell(row=current_row, column=2, value=to_val)
                    ws.cell(row=current_row, column=3, value=via_val)
                    ws.cell(row=current_row, column=4, value=sub["desc"])
                    ws.cell(row=current_row, column=5, value=sub["curr"])
                    ws.cell(row=current_row, column=6, value=sub["20"])
                    ws.cell(row=current_row, column=7, value=sub["40"])
                    ws.cell(row=current_row, column=8, value=sub["40hc"])
                    ws.cell(row=current_row, column=9, value="")
                    current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20   # From
    ws.column_dimensions['B'].width = 25   # To
    ws.column_dimensions['C'].width = 20   # Via
    ws.column_dimensions['D'].width = 50   # Description
    ws.column_dimensions['E'].width = 10   # Curr.
    ws.column_dimensions['F'].width = 12   # 20STD
    ws.column_dimensions['G'].width = 12   # 40STD
    ws.column_dimensions['H'].width = 12   # 40HC
    ws.column_dimensions['I'].width = 40   # Transport Remarks
    
    # Save
    output_path = os.path.join("downloads", "hapag_surcharges.xlsx")
    wb.save(output_path)
    print(f"Created mock HAPAG surcharges file: {output_path}")
    print(f"  Destinations: {len(destinations_data)}")
    print(f"  Total rows: {current_row - 5}")
    
    return output_path

if __name__ == "__main__":
    create_mock_hapag_data()
