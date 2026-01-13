"""
Hapag-Lloyd Quote Checker - Automated Price Extraction

This script automates the process of:
1. Logging into Hapag-Lloyd
2. Searching for shipping quotes between ports
3. Selecting offers and viewing price breakdowns
4. Extracting Import Surcharges table to Excel
"""

import os
import json
import time
from datetime import datetime
from playwright.sync_api import Playwright, sync_playwright, expect
from playwright_stealth import Stealth
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# Load environment variables
load_dotenv()


def load_destinations():
    """Load destinations from destinations.txt"""
    destinations = []
    with open("destinations.txt", "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                destinations.append(line)
    return destinations


def load_destination_configs():
    """Load destination configurations from destination_configs.json"""
    with open("destination_configs.json", "r", encoding="utf-8") as f:
        return json.load(f)


def determine_excel_filename(base_filename="hapag_surcharges.xlsx"):
    """
    Determine the Excel filename to use for this run.
    Checks once at the start if base file exists, creates dated version if needed.
    
    Returns:
        str: Filename to use (e.g., "hapag_surcharges.xlsx" or "hapag_surcharges_20260113_2.xlsx")
    """
    downloads_dir = os.path.join(os.getcwd(), "downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    
    base_path = os.path.join(downloads_dir, base_filename)
    
    if not os.path.exists(base_path):
        return base_filename
    
    # Create dated filename
    today = datetime.now().strftime("%Y%m%d")
    base_name = base_filename.rsplit('.', 1)[0]
    ext = base_filename.rsplit('.', 1)[1] if '.' in base_filename else 'xlsx'
    
    # Check if dated file exists, add _2, _3, etc.
    counter = 1
    dated_filename = f"{base_name}_{today}.{ext}"
    dated_path = os.path.join(downloads_dir, dated_filename)
    
    while os.path.exists(dated_path):
        counter += 1
        dated_filename = f"{base_name}_{today}_{counter}.{ext}"
        dated_path = os.path.join(downloads_dir, dated_filename)
    
    return dated_filename


def extract_route_info(page):
    """
    Extract route information (From, To, Via) from Price Breakdown dialog
    
    Returns:
        dict: {"from": "BUSAN", "to": "VALENCE/DROME", "via": "FOS SUR MER"}
    """
    route = {"from": "", "to": "", "via": ""}
    
    try:
        # Extract "From" location
        from_text = page.get_by_text("From", exact=True).locator("..").inner_text()
        # Parse to get just the location name (skip "From" label)
        route["from"] = from_text.replace("From", "").strip()
    except:
        print("   [WARNING] Could not extract 'From' location")
    
    try:
        # Extract "To" location
        to_text = page.get_by_text("To", exact=True).locator("..").inner_text()
        route["to"] = to_text.replace("To", "").strip()
    except:
        print("   [WARNING] Could not extract 'To' location")
    
    try:
        # Extract "via" location
        via_text = page.get_by_text("via").locator("..").inner_text()
        route["via"] = via_text.replace("via", "").strip()
    except:
        print("   [WARNING] Could not extract 'via' location")
    
    print(f"   [ROUTE] From: {route['from']}, To: {route['to']}, Via: {route['via']}")
    return route


def run(playwright: Playwright) -> None:
    """Main automation workflow for Hapag-Lloyd quote checking"""
    
    # Load destinations and configs
    destinations = load_destinations()
    configs = load_destination_configs()
    
    # Get credentials from environment variables
    email = os.getenv('HAPAG_EMAIL')
    password = os.getenv('HAPAG_PASSWORD')
    
    if not email or not password:
        print("❌ ERROR: HAPAG_EMAIL and HAPAG_PASSWORD must be set in .env file")
        return
    
    # Determine Excel file path (check once at start)
    excel_filename = determine_excel_filename()
    print(f"📁 Excel file for this run: {excel_filename}\n")
    
    print(f"\n{'='*60}")
    print(f"📋 Processing {len(destinations)} destinations from destinations.txt")
    print(f"{'='*60}")
    for dest in destinations:
        print(f"  • {dest}")
    print(f"{'='*60}\n")
    
    # Launch browser (headless=False to see the browser)
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    
    # Apply stealth mode to avoid detection
    stealth = Stealth()
    stealth.apply_stealth_sync(page)
    
    # Navigate to Hapag-Lloyd quote page
    print("🌐 Navigating to Hapag-Lloyd...")
    page.goto("https://www.hapag-lloyd.com/solutions/new-quote/#/simple?language=en")
    
    # Wait for login fields to appear (skip Cloudflare - user will handle manually)
    print("⏳ Waiting for login page to appear (handle Cloudflare manually if needed)...")
    page.get_by_role("textbox", name="E-mail Address").wait_for(timeout=120000)  # 2 min timeout
    time.sleep(1)
    
    # Login (only once)
    print("🔐 Logging in...")
    try:
        page.get_by_role("button", name="Select All").click()
        time.sleep(0.5)
    except:
        print("   [INFO] Cookie consent already accepted")
    
    page.get_by_role("textbox", name="E-mail Address").click()
    page.get_by_role("textbox", name="E-mail Address").fill(email)
    page.get_by_role("textbox", name="E-mail Address").press("Tab")
    page.get_by_role("textbox", name="Password").fill(password)
    page.get_by_role("textbox", name="Password").press("Enter")
    
    # Wait for redirect back to quote page
    print("⏳ Waiting for quote page to load after login...")
    page.get_by_test_id("start-input").wait_for(timeout=30000)
    time.sleep(2)
    
    # Process each destination
    for idx, destination in enumerate(destinations, 1):
        print(f"\n{'='*60}")
        print(f"🎯 Processing {idx}/{len(destinations)}: {destination}")
        print(f"{'='*60}")
        
        # Get location code from config
        if destination not in configs:
            print(f"❌ ERROR: '{destination}' not found in destination_configs.json")
            continue
        
        location_code = configs[destination]["locationCode"]
        print(f"📍 Location Code: {location_code}")
        
        # If not first destination, click Edit to start new search
        if idx > 1:
            print("✏️ Starting new search...")
            page.get_by_role("button", name="Edit").first.click()
            page.get_by_role("listitem").filter(has_text="Edit Search").click()
            time.sleep(1)
        
        # Enter origin port (BUSAN) - only for first destination
        if idx == 1:
            print("📍 Entering origin: BUSAN (KRPUS)...")
            page.get_by_test_id("start-input").click()
            page.get_by_test_id("start-input").fill("busan")
            time.sleep(2.5)  # Wait for autocomplete dropdown to populate
            # Select the correct Busan - KRPUS (Korea), not ITBSN (Italy)
            try:
                page.get_by_text("BUSAN (KRPUS)").click()
                time.sleep(1)
            except:
                print("   ⚠️ Could not find BUSAN (KRPUS), using arrow key")
                page.get_by_test_id("start-input").press("ArrowDown")
                page.get_by_test_id("start-input").press("Enter")
                time.sleep(1)
        else:
            print("   ℹ️ Keeping origin: BUSAN (KRPUS) from previous search")
        
        # Clear destination if needed
        if idx > 1:
            try:
                page.get_by_test_id("end-column").get_by_role("button", name="Clear").click()
                time.sleep(0.5)
            except:
                pass
        
        # Enter destination using location code
        print(f"📍 Entering destination: {location_code}...")
        page.get_by_test_id("end-input").click()
        page.get_by_test_id("end-input").fill(location_code.lower())
        time.sleep(1.5)  # Wait for autocomplete dropdown to populate
        
        # Try to click the exact match with location code
        try:
            page.get_by_text(f"({location_code})").first.click()
            time.sleep(0.5)
        except:
            print(f"⚠️ WARNING: Could not find exact match for {location_code}, using arrow key selection")
            page.get_by_test_id("end-input").press("ArrowDown")
            page.get_by_test_id("end-input").press("Enter")
            time.sleep(0.5)
        
        # Select delivery option
        print("🚚 Selecting 'Delivered to your Door'...")
        time.sleep(0.5)
        page.get_by_role("radio", name="Delivered to your Door (").click()
        time.sleep(0.5)
        
        # Search for quotes
        print("🔍 Searching for quotes...")
        page.get_by_test_id("search-submit").click()
        
        # Wait for search results to load - use smart waiting for Price Breakdown button
        print("⏳ Waiting for search results and Price Breakdown button...")
        try:
            # Wait for Price Breakdown button to be visible (up to 30 seconds)
            page.get_by_role("button", name="Price Breakdown").first.wait_for(state="visible", timeout=30000)
            print("   ✅ Price Breakdown button found!")
            time.sleep(1)  # Brief pause to ensure page is stable
        except Exception as e:
            print(f"❌ ERROR: Price Breakdown button did not appear: {e}")
            continue
        
        # Click Price Breakdown button
        print("💰 Opening Price Breakdown...")
        try:
            page.get_by_role("button", name="Price Breakdown").first.click()
            time.sleep(3)  # Wait for dialog to fully load
        except Exception as e:
            print(f"❌ ERROR: Could not click Price Breakdown: {e}")
            continue
        
        # Extract route information (From, To, Via)
        print("📍 Extracting route information...")
        route_info = extract_route_info(page)
        
        # Extract Import Surcharges table
        print("📊 Extracting Import Surcharges table...")
        table_data = extract_import_surcharges_table(page)
        
        if not table_data:
            print(f"⚠️ WARNING: No data extracted for {destination}")
            # Close dialog and continue
            try:
                page.keyboard.press("Escape")
                time.sleep(1)
            except:
                pass
            continue
        
        # Save to Excel (append mode) with route info
        print("💾 Saving to Excel...")
        excel_path = save_to_excel(table_data, destination=destination, route_info=route_info, filename=excel_filename)
        print(f"✅ Data saved to: {excel_path}")
        
        # Close the Price Breakdown dialog
        try:
            page.keyboard.press("Escape")
            time.sleep(1)
        except:
            pass
        
        print(f"✅ Completed {destination}")
    
    # Summary
    print(f"\n{'='*60}")
    print(f"🎉 All {len(destinations)} destinations processed!")
    print(f"{'='*60}")
    print(f"📁 Results saved to: downloads/hapag_surcharges.xlsx")
    
    # Keep browser open for review
    input("\nPress Enter to close browser...")
    
    # Cleanup
    context.close()
    browser.close()
    print("✅ Automation complete!")


def extract_import_surcharges_table(page):
    """
    Extract Import Surcharges table data using dynamic header parsing.
    Works with any combination of container types (20STD, 40STD, 40HC, etc.)
    
    Returns:
        list: List of dicts with keys: description, curr, container_values, remarks
              container_values is a dict like {"40STD": "100", "40HC": "120"}
    """
    print("   [INFO] Extracting Import Surcharges table using dynamic header parsing...")
    
    import re
    
    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip())

    def split_first_cell(text: str):
        lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
        if not lines:
            return "", ""
        if len(lines) == 1:
            return lines[0], ""
        return lines[0], " ".join(lines[1:])

    # Wait until at least one known row is present
    try:
        page.get_by_role("cell", name="Terminal Handling Charge Dest.").wait_for(timeout=30000)
        print("   [INFO] Table loaded successfully")
    except Exception as e:
        print(f"   [WARNING] Timeout waiting for table: {e}")

    # Find all rows
    rows = page.get_by_role("row")
    row_count = rows.count()
    print(f"   [INFO] Found {row_count} rows")
    
    if row_count == 0:
        rows = page.locator("tr")
        row_count = rows.count()
        print(f"   [INFO] Fallback: Found {row_count} <tr> rows")

    # Step 1: Find header row and build column map
    header_map = {}
    header_row_index = -1
    
    for i in range(min(row_count, 20)):  # Check first 20 rows for header
        r = rows.nth(i)
        cells = r.get_by_role("cell")
        if cells.count() < 3:
            continue
        
        # Check if this looks like a header row
        cell_texts = [norm(cells.nth(j).inner_text()) for j in range(cells.count())]
        
        # Look for currency column as indicator
        if "Curr." in cell_texts or "Currency" in cell_texts:
            header_row_index = i
            for j, text in enumerate(cell_texts):
                header_map[j] = text
            print(f"   [INFO] Found header row at index {i}")
            print(f"   [INFO] Header columns: {header_map}")
            break
    
    if not header_map:
        print("   [WARNING] No header row found, using position-based extraction")
        # Fallback: assume standard structure
        header_map = {0: "Description", 1: "Curr.", 2: "40STD", 3: "40HC"}
    
    # Step 2: Identify container type columns (20STD, 40STD, 40HC, etc.)
    container_columns = {}
    for idx, col_name in header_map.items():
        # Match pattern like "20STD", "40STD", "40HC", "45HC"
        if re.match(r'\d+[A-Z]+', col_name):
            container_columns[idx] = col_name
    
    print(f"   [INFO] Container columns found: {container_columns}")
    
    # Step 3: Find currency column index and description column
    curr_idx = None
    desc_idx = 0  # Usually first column
    for idx, col_name in header_map.items():
        if "Curr" in col_name or col_name == "Currency":
            curr_idx = idx
            break
    
    # Step 4: Extract data rows
    data = []
    extracted_count = 0
    
    for i in range(row_count):
        if i == header_row_index:
            continue  # Skip header row
            
        r = rows.nth(i)
        cells = r.get_by_role("cell")
        cell_count = cells.count()
        
        if cell_count < 3:
            continue

        first = cells.nth(desc_idx).inner_text()
        desc, remarks = split_first_cell(first)

        # Filter out header-like rows
        if desc in ["Import Surcharges", "Export Surcharges", "Description", "Curr.", "40STD", "40HC", "20STD", ""]:
            continue

        # Extract currency
        curr = norm(cells.nth(curr_idx).inner_text()) if curr_idx is not None and curr_idx < cell_count else ""
        
        # Extract container values dynamically - use the EXACT column indices from header_map
        container_values = {}
        for col_idx, col_name in sorted(container_columns.items()):
            if col_idx < cell_count:
                value = norm(cells.nth(col_idx).inner_text())
                container_values[col_name] = value
                print(f"      [COL {col_idx}] {col_name} = {value}")
        
        if desc:
            data.append({
                "description": desc,
                "curr": curr,
                "container_values": container_values,
                "remarks": remarks,
                "has_20std_column": "20STD" in container_columns.values()  # Track if source has 20STD column
            })
            extracted_count += 1
            print(f"   [EXTRACTED] {desc}: {curr} {container_values} | Remarks: {remarks or 'N/A'}")

    if not data:
        print("   [ERROR] No data successfully extracted")
    else:
        print(f"   [SUCCESS] ✅ Extracted {extracted_count} rows from Import Surcharges")
    
    return data


def save_to_excel(data, destination="", route_info=None, filename="hapag_surcharges.xlsx"):
    """
    Save extracted data to Excel file with smart column mapping.
    - Fixed columns: From, To, Via, Description, Curr., 20STD, 40STD, 40HC, Transport Remarks
    - If 20STD not in data, copy 40STD value
    - Appends to existing file if present
    
    Args:
        data: List of dicts with keys: description, curr, container_values, remarks
        destination: Name of the destination city/country
        route_info: Dict with keys: from, to, via
        filename: Name of the Excel file
        
    Returns:
        str: Path to saved Excel file
    """
    if route_info is None:
        route_info = {"from": "", "to": "", "via": ""}
    
    # Create downloads directory if it doesn't exist
    downloads_dir = os.path.join(os.getcwd(), "downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    
    excel_path = os.path.join(downloads_dir, filename)
    
    # Fixed column structure
    fixed_headers = ["From", "To", "Via", "Description", "Curr.", "20STD", "40STD", "40HC", "Transport Remarks"]
    
    # Check if file exists for append mode
    if os.path.exists(excel_path):
        print(f"   [INFO] Appending to existing file: {excel_path}")
        wb = load_workbook(excel_path)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        print(f"   [INFO] Creating new file: {excel_path}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Surcharges"
        
        # Add metadata
        ws["H1"] = "Generated:"
        ws["I1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["H2"] = "Origin:"
        ws["I2"] = "BUSAN (KRPUS)"
        
        # Write headers
        for col_idx, header in enumerate(fixed_headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        start_row = 5
    
    # Write data rows
    for row_data in data:
        desc = row_data["description"]
        curr = row_data["curr"]
        containers = row_data["container_values"]
        remarks = row_data["remarks"]
        has_20std_column = row_data.get("has_20std_column", False)
        
        # Smart column mapping with 40STD→20STD fallback
        val_20std = containers.get("20STD", "")
        val_40std = containers.get("40STD", "")
        val_40hc = containers.get("40HC", "")
        
        # Only copy 40STD to 20STD if source table doesn't have 20STD column at all
        if not has_20std_column and val_40std:
            val_20std = val_40std
            print(f"   [MAPPING] Source table has no 20STD column, copying 40STD ({val_40std}) → 20STD for '{desc}'")
        
        # Map to fixed columns (From, To, Via, Description, Curr., 20STD, 40STD, 40HC, Transport Remarks)
        row_values = [
            route_info["from"],
            route_info["to"],
            route_info["via"],
            desc,
            curr,
            val_20std,
            val_40std,
            val_40hc,
            remarks
        ]
        
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=start_row, column=col_idx, value=value)
        
        start_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # From
    ws.column_dimensions['B'].width = 25  # To
    ws.column_dimensions['C'].width = 20  # Via
    ws.column_dimensions['D'].width = 40  # Description
    ws.column_dimensions['E'].width = 10  # Curr.
    ws.column_dimensions['F'].width = 12  # 20STD
    ws.column_dimensions['G'].width = 12  # 40STD
    ws.column_dimensions['H'].width = 12  # 40HC
    ws.column_dimensions['I'].width = 50  # Transport Remarks
    
    # Save workbook
    wb.save(excel_path)
    
    return excel_path
    ws.column_dimensions['A'].width = 30  # Destination
    ws.column_dimensions['B'].width = 40  # Description
    ws.column_dimensions['C'].width = 10  # Curr.
    ws.column_dimensions['D'].width = 12  # 20STD
    ws.column_dimensions['E'].width = 12  # 40STD
    ws.column_dimensions['F'].width = 12  # 40HC
    ws.column_dimensions['G'].width = 50  # Transport Remarks
    
    # Save workbook
    wb.save(excel_path)
    
    return excel_path


if __name__ == "__main__":
    with sync_playwright() as playwright:
        run(playwright)
